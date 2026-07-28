# -*- coding: utf-8 -*-
"""Rebuild bộ data NER sạch từ Excel annotation gốc (NER_AnhThieu).

Nguồn: các file .xlsx trong --excel-dir, 2 dạng:
  - Series NER-1-100 ... NER-901-1000: cột ``chinese`` chứa inline markup
    ``"entity"(TYPE)`` — parse trực tiếp ra span (tin cậy cao).
  - Series NER 1288-*: file TÍCH LŨY (file sau chứa cả file trước), cột
    ``chinese`` không markup, cột ``NER`` là dict {'PER': [...], ...} —
    project entity lên text bằng exact-match trong phạm vi row.

Xử lý các lỗi đã biết của nguồn: nhãn typo (TITTLE, DATE, og, PERr...),
orphan markup ``"(TYPE)`` do find-replace ẩu, dòng trùng do file tích lũy.

Output (--out-dir):
  - {train,dev,test}.jsonl  — format instruction/input/output như dataset cũ
  - {train,dev,test}.bio.txt — BIO 1 ký tự/dòng
  - report.json             — thống kê + danh sách row bị loại
Split theo row unique, KHÔNG có leakage (dedup tuyệt đối trước khi split).
"""
import argparse
import ast
import glob
import json
import os
import random
import re
import sys
from collections import Counter

sys.stdout.reconfigure(encoding="utf-8", errors="replace")

MARK = re.compile(r'"([^"]{1,40})"\(([^)]{1,25})\)')
# markup mồ côi ngay sau 1 annotation vừa đóng: ..."(TITLE)"(TITLE)
ORPHAN_AFTER = re.compile(r'(?<=\))"\((?:PER|LOC|ORG|DTM|DATE|TITLE|TITTLE)\)')
# mảnh vỡ do quote lệch còn lại trong đoạn text thường: "(ORG) hoặc " đơn lẻ
FRAG = re.compile(r'"\([A-Za-z. ]{1,15}\)|"')
PUNCT = set('.,、。；:：!?！？"()（）[]')

TYPE_MAP = {
    "PER": "PER", "LOC": "LOC", "ORG": "ORG", "TITLE": "TITLE",
    "DATE": "DTM", "DTM": "DTM", "TITTLE": "TITLE",
    "TITLE PER": "TITLE", "TITLEPER": "TITLE",
    "og": "ORG", "o": "ORG", "OR": "ORG", "ORGg": "ORG", "PERr": "PER",
    "ORG.": "ORG",
    # ngoài schema 5 nhãn -> bỏ (giữ text, không giữ nhãn)
    "NAME OF BOOK": None, "NAMEOFBOOK": None, "BOOK": None, "book": None,
}

INSTRUCTION = (
    "你是一个专门处理越南汉文古籍的命名实体识别系统。请识别以下文本中的命名实体，"
    "并用{实体|类型}格式标注。\n实体类型：\n- PER: 人名 (人物姓名)\n"
    "- LOC: 地名 (地理位置)  \n- ORG: 机构名 (组织机构)\n- DTM: 时间 (日期时间)\n"
    "- TITLE: 官职 (官职称号)\n只输出标注后的文本，不要解释。"
)


def clean_text(t: str) -> str:
    """Chuẩn hoá text cell: bỏ newline/space (dataset cũ cũng không có)."""
    return t.replace("\n", "").replace(" ", "").replace("\u3000", "")


def parse_marked(text: str):
    """Parse inline markup -> (plain, spans, err). spans = [(start, end, type)].

    Thứ tự quan trọng: bỏ orphan ``)"(TYPE)`` TRƯỚC (lookbehind nên không đụng
    annotation hợp lệ), parse MARK, rồi mới dọn mảnh vỡ quote lệch trong các
    đoạn text thường. Dòng có quá nhiều mảnh vỡ -> flag để loại.
    """
    text = ORPHAN_AFTER.sub("", text)
    plain_parts, spans = [], []
    pos = 0
    length = 0
    n_frag = 0
    for m in MARK.finditer(text):
        seg, nf = FRAG.subn("", text[pos:m.start()])
        n_frag += nf
        plain_parts.append(seg)
        length += len(seg)
        surf, typ = m.group(1), m.group(2).strip()
        mapped = TYPE_MAP.get(typ, "__UNKNOWN__")
        if mapped == "__UNKNOWN__":
            # nhãn lạ (thường do quote hỏng nặng) -> flag cả row
            return None, None, f"unknown type {typ!r}"
        plain_parts.append(surf)
        if mapped is not None:
            spans.append((length, length + len(surf), mapped))
        length += len(surf)
        pos = m.end()
    tail, nf = FRAG.subn("", text[pos:])
    n_frag += nf
    plain_parts.append(tail)
    plain = "".join(plain_parts)
    if n_frag > 3:
        return None, None, f"too many broken fragments ({n_frag})"
    if '"' in plain or "(" in plain or ")" in plain:
        return None, None, "residual markup after parse"
    return plain, spans, None


def parse_ner_dict(s: str):
    """Parse chuỗi dict {'PER': [...]} (có thể hỏng nhẹ) -> {type: [surface]}."""
    if not s or s in ("None", ""):
        return {}
    try:
        d = ast.literal_eval(s)
    except (ValueError, SyntaxError):
        # dict bị cắt cụt -> vá ngoặc rồi thử lại
        for suffix in ("']}", "]}", "}"):
            try:
                d = ast.literal_eval(s + suffix)
                break
            except (ValueError, SyntaxError):
                d = None
        if d is None:
            return {}
    if not isinstance(d, dict):
        return {}
    out = {}
    for k, v in d.items():
        mapped = TYPE_MAP.get(str(k).strip())
        if mapped is None or not isinstance(v, (list, tuple)):
            continue
        surfs = []
        for e in v:
            e = str(e).strip().strip('.。"\'` ')
            e = clean_text(e)
            if len(e) >= 2 and not any(c in PUNCT for c in e):
                surfs.append(e)
        if surfs:
            out.setdefault(mapped, []).extend(surfs)
    return out


def project_dict(plain: str, ner: dict):
    """Project entity không vị trí lên text: exact-match mọi lần xuất hiện,
    ưu tiên surface dài trước để tránh chồng lấn."""
    taken = [False] * len(plain)
    spans = []
    items = [(s, t) for t, ss in ner.items() for s in ss]
    items.sort(key=lambda x: -len(x[0]))
    for surf, typ in items:
        start = 0
        while True:
            p = plain.find(surf, start)
            if p < 0:
                break
            start = p + 1
            if any(taken[p:p + len(surf)]):
                continue
            spans.append((p, p + len(surf), typ))
            for i in range(p, p + len(surf)):
                taken[i] = True
    spans.sort()
    return spans


def spans_to_inline(plain: str, spans):
    out, pos = [], 0
    for st, en, tp in spans:
        out.append(plain[pos:st])
        out.append("{" + plain[st:en] + "|" + tp + "}")
        pos = en
    out.append(plain[pos:])
    return "".join(out)


def spans_to_bio(plain: str, spans):
    labs = ["O"] * len(plain)
    for st, en, tp in spans:
        labs[st] = "B-" + tp
        for i in range(st + 1, en):
            labs[i] = "I-" + tp
    return labs


def load_rows(excel_dir: str):
    """Đọc mọi .xlsx, dedup theo plain text, ưu tiên bản có inline markup."""
    import openpyxl
    rows, order = {}, []
    for f in sorted(glob.glob(os.path.join(excel_dir, "*.xlsx"))):
        wb = openpyxl.load_workbook(f, read_only=True)
        ws = wb.worksheets[0]
        hdr = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
        if "chinese" not in hdr:
            wb.close()
            continue
        ci = hdr.index("chinese")
        ni = hdr.index("NER") if "NER" in hdr else None
        for row in ws.iter_rows(min_row=2, values_only=True):
            ch = row[ci]
            if not ch or not isinstance(ch, str) or not ch.strip():
                continue
            ner = row[ni] if ni is not None and ni < len(row) else None
            plain_key = clean_text(FRAG.sub("", MARK.sub(
                lambda m: m.group(1), ORPHAN_AFTER.sub("", ch))))
            marked = bool(MARK.search(ch))
            cur = rows.get(plain_key)
            if cur is None:
                rows[plain_key] = [ch, marked, ner, os.path.basename(f)]
                order.append(plain_key)
            else:
                if marked and not cur[1]:
                    cur[0], cur[1] = ch, True
                if ner and not cur[2]:
                    cur[2] = ner
        wb.close()
    return [(k, *rows[k]) for k in order]


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--excel-dir", default=r"D:\bio_source\NER_zip\NER\NER_AnhThieu")
    ap.add_argument("--out-dir", default=r"D:\ancient-chinese-ner\data\processed\ner_clean")
    ap.add_argument("--seed", type=int, default=42)
    ap.add_argument("--min-len", type=int, default=10)
    args = ap.parse_args()

    raw = load_rows(args.excel_dir)
    print(f"unique rows: {len(raw)}")

    records, dropped = [], []
    n_inline, n_dict = 0, 0
    for key, ch, marked, ner, src in raw:
        if marked:
            plain, spans, err = parse_marked(clean_text(ch))
            if err:
                dropped.append({"src": src, "reason": err, "text": clean_text(ch)[:80]})
                continue
            source = "inline"
            n_inline += 1
        else:
            plain = clean_text(ch)
            nd = parse_ner_dict(str(ner)) if ner else {}
            if not nd:
                dropped.append({"src": src, "reason": "no annotation", "text": plain[:80]})
                continue
            spans = project_dict(plain, nd)
            source = "dict"
            n_dict += 1
        if len(plain) < args.min_len or not spans:
            dropped.append({"src": src, "reason": "too short / no entities", "text": plain[:80]})
            continue
        # validator: entity không chứa dấu câu
        bad = [plain[st:en] for st, en, _ in spans if any(c in PUNCT for c in plain[st:en])]
        if bad:
            dropped.append({"src": src, "reason": f"punct in entity {bad[:3]}", "text": plain[:80]})
            continue
        records.append({"plain": plain, "spans": spans, "source": source})

    print(f"kept: {len(records)} (inline {n_inline}, dict {n_dict}) | dropped: {len(dropped)}")

    # dedup lần cuối + split không leakage
    seen = set()
    uniq = []
    for r in records:
        if r["plain"] in seen:
            continue
        seen.add(r["plain"])
        uniq.append(r)
    random.seed(args.seed)
    random.shuffle(uniq)
    n = len(uniq)
    n_test = n_dev = max(1, n // 10)
    splits = {
        "test": uniq[:n_test],
        "dev": uniq[n_test:n_test + n_dev],
        "train": uniq[n_test + n_dev:],
    }

    os.makedirs(args.out_dir, exist_ok=True)
    type_dist = Counter()
    for name, recs in splits.items():
        with open(os.path.join(args.out_dir, f"{name}.jsonl"), "w", encoding="utf-8") as fj, \
             open(os.path.join(args.out_dir, f"{name}.bio.txt"), "w", encoding="utf-8") as fb:
            for r in recs:
                fj.write(json.dumps({
                    "instruction": INSTRUCTION,
                    "input": r["plain"],
                    "output": spans_to_inline(r["plain"], r["spans"]),
                }, ensure_ascii=False) + "\n")
                labs = spans_to_bio(r["plain"], r["spans"])
                for c, l in zip(r["plain"], labs):
                    fb.write(f"{c} {l}\n")
                fb.write("\n")
                for _, _, tp in r["spans"]:
                    type_dist[tp] += 1
        print(f"{name}: {len(recs)} records")

    n_ents = sum(len(r["spans"]) for r in uniq)
    report = {
        "unique_rows_source": len(raw),
        "kept": len(uniq),
        "from_inline_markup": n_inline,
        "from_dict_projection": n_dict,
        "dropped": len(dropped),
        "total_entities": n_ents,
        "type_distribution": dict(type_dist),
        "splits": {k: len(v) for k, v in splits.items()},
        "dropped_detail": dropped,
    }
    with open(os.path.join(args.out_dir, "report.json"), "w", encoding="utf-8") as f:
        json.dump(report, f, ensure_ascii=False, indent=2)
    print(f"total entities: {n_ents} | type dist: {dict(type_dist)}")
    print(f"report -> {os.path.join(args.out_dir, 'report.json')}")


if __name__ == "__main__":
    main()
