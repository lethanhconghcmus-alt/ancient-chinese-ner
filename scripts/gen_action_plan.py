#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""Sinh file kế hoạch xử lý tồn đọng do dữ liệu v1 hỏng -> ke_hoach_VCL2026.xlsx"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

OUT = r"D:\ancient-chinese-ner\ke_hoach_VCL2026.xlsx"

H = PatternFill("solid", fgColor="1F3864")
HF = Font(bold=True, color="FFFFFF", size=10)
BOLD = Font(bold=True, size=10)
BASE = Font(size=10)
WRAP = Alignment(wrap_text=True, vertical="top")
TOP = Alignment(vertical="top")
CEN = Alignment(horizontal="center", vertical="top")
THIN = Border(*[Side(style="thin", color="BFBFBF")] * 4)

P1 = PatternFill("solid", fgColor="FFC7CE")   # chặn đường găng
P2 = PatternFill("solid", fgColor="FFEB9C")   # quan trọng
P3 = PatternFill("solid", fgColor="D9E1F2")   # nên làm
P4 = PatternFill("solid", fgColor="E2EFDA")   # hoãn được
OK = PatternFill("solid", fgColor="C6EFCE")
BAD = PatternFill("solid", fgColor="FFC7CE")

PF = {"P0": P1, "P1": P2, "P2": P3, "P3": P4}


def sheet(wb, title, headers, rows, widths, first=False):
    ws = wb.active if first else wb.create_sheet()
    ws.title = title
    ws.append(headers)
    for c in range(1, len(headers) + 1):
        cell = ws.cell(1, c)
        cell.fill, cell.font, cell.alignment, cell.border = H, HF, CEN, THIN
    for r in rows:
        ws.append(r)
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.font, cell.alignment, cell.border = BASE, WRAP, THIN
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{ws.max_row}"
    return ws


wb = openpyxl.Workbook()

# ══════════════════════════════════════════════════ 1. KẾ HOẠCH
HDR = ["ID", "Ưu tiên", "Tuần", "Nhóm", "Việc cần làm", "Vì sao / Phụ thuộc",
       "OUTPUT cụ thể (file / con số)", "Ước lượng", "Trạng thái", "Ghi chú"]

TASKS = [
 # ---------- P0: chặn đường găng ----------
 ["A1","P0","T1 (04–10/08)","Data",
  "Chạy clean_labels.py lên TEST",
  "BẮT BUỘC. Train/dev đã áp guideline v1.0, test chưa → model bị trừ điểm vì trả lời ĐÚNG. "
  "Không xung đột gold review (áp sau vẫn được).",
  "test.jsonl + test.bio.txt đã chuẩn hoá; clean_report cập nhật; "
  "nhiễu test 2.76% → kỳ vọng <0.5%","30 phút","Chưa",
  "Gỡ chặn toàn bộ. KHÔNG cần duyệt xong 924 dòng mới rerun được."],

 ["A2","P0","T1","Data",
  "Xử lý 6 record dev bị leakage",
  "6/160 dev (idx 29,30,69,83,99,117) chồng lấn train — đoạn Chiêm Thành 49 ký tự duy nhất trong "
  "nguyên bản. Làm lệch việc chọn checkpoint. Đề xuất: BỎ 7 record train (285,601,640,642,729,1062,1144).",
  "train.jsonl còn 1,277 record; dev giữ nguyên 160 để so được số cũ; "
  "ghi log vào clean_report.json","30 phút","Chưa",
  "Bỏ bên train an toàn hơn bỏ bên dev."],

 ["A3","P0","T1","Backup",
  "Đưa artifact gold review vào git",
  "gold_review_sheet.xlsx (58 dòng đã duyệt) nằm NGOÀI git, không backup. "
  "review_rules.py ĐÃ MẤT khỏi máy — chỉ còn output json.",
  "data/review/{gold_review_sheet.xlsx, review_proposals_rules.json, manual_batch1-4.json} "
  "đã commit","20 phút","Chưa",
  "Rủi ro sống còn duy nhất còn lại. Làm trước khi duyệt tiếp."],

 ["A4","P0","T1","Baseline",
  "Rerun GuwenBERT-CRF trên v2-clean",
  "Thay số 82.13% của Paper 1. Phụ thuộc A1+A2. "
  "Repo chisiec-ner đọc CoNLL từ data/processed/ner_clean/*.bio.txt.",
  "F1 tổng + per-entity (PER/LOC/ORG/DTM/TITLE) trên v2-clean; "
  "results/baseline_v2/metrics.json","3–4 giờ","Chưa",
  "Con số THẬT đầu tiên. Dự kiến tụt mạnh so với 82.13 — đó là điều đúng."],

 ["A5","P0","T2 (11–17/08)","Paper 2",
  "Rerun E2 (pretrain+SFT) trên v2-clean",
  "Thay số 0.7507. Phụ thuộc A1+A2. Pipeline 3 notebook đã chạy được (~4.5h).",
  "E2 F1 trên v2-clean; results/E2_v2/; so sánh cạnh E2 cũ (v1)","5 giờ","Chưa",
  "Kết quả chính của Paper 2."],

 # ---------- P1: quan trọng ----------
 ["B1","P1","T1","Viết",
  "Rà lại ABSTRACT đã đăng ký",
  "Chưa submit full paper nhưng abstract đã có, một phần dùng kết quả v1. "
  "Nguy hiểm nhất là nếu abstract quảng cáo quy mô corpus.",
  "Bản abstract đã sửa; xác nhận venue có cho sửa trước 30/08 không","2 giờ","Chưa",
  "7,301 câu / 107,105 entity TỰ THÂN đã sai (nhân bản 3.7×). Thật: 1,604 record."],

 ["B2","P1","T2","Data",
  "Tính lại thống kê corpus cho paper",
  "Mọi con số mô tả dữ liệu trong paper đang là của v1.",
  "Bảng: #record, #token, #entity, phân bố 5 type, split 1284/160/160, "
  "%phủ DVSKTT (71.6%), #section vắng","1 giờ","Chưa",
  "Lấy từ clean_report.json + record_source_map.json."],

 ["B3","P1","T2","Paper 2",
  "Sinh lại error analysis trên v2",
  "Mục 6.4 Paper 1 viết từ error_analysis.txt của v1. Phân tích lỗi cũ đang "
  "mô tả lỗi của memorization, không phải lỗi NER thật.",
  "error_analysis_v2.txt + bảng confusion 5×5; phân tích TITLE↔ORG "
  "(vốn là lỗi lớn nhất)","2 giờ","Chưa",
  "Có thể phân tích theo THỜI KỲ nhờ record_source_map.json — điểm mới."],

 ["B4","P1","T1–T3","Data",
  "Tiếp tục duyệt gold test (924 dòng)",
  "ĐANG LÀM, 58/982 (5.9%). KHÔNG còn gating sau khi có A1.",
  "gold_review_sheet.xlsx cột DUYỆT đầy đủ; script apply; test-gold v2.1",
  "~15 giờ","Đang làm",
  "86% đề xuất được OK thẳng → batch-accept nhóm căn cứ chắc, chỉ soi tay RETYPE/BOUNDARY."],

 ["B5","P1","T3 (18–24/08)","Viết",
  "Viết Paper 2 theo khung mới",
  "Đổi trọng tâm đóng góp: từ 'đạt F1 cao' sang 'phát hiện + sửa lỗi corpus + "
  "benchmark hiệu chỉnh'. Contribution mạnh hơn cho miền ít tài nguyên.",
  "Bản nháp full paper; bảng so sánh v1 vs v2 làm bằng chứng","20 giờ","Chưa",
  "Vật liệu đã có sẵn ở docs/dataset_v2_cleaning.md."],

 # ---------- P2: nên làm ----------
 ["C1","P2","T2","Data",
  "Quyết 216 chuỗi needs_review",
  "Ambiguity thật, cần guideline v1.1: 哀牢 ORG/LOC · 蠻 PER/ORG · 李氏 ORG/PER · 大行 PER/TITLE.",
  "guideline v1.1 + rerun clean_labels.py","3 giờ","Chưa",
  "Gom về ~10–15 quyết định, không phải 216 phán đoán."],

 ["C2","P2","T2","Data",
  "Xử lý noise văn bản còn lại",
  "Folio marker Latin (rec#92,#143 dev + train) → XOÁ. "
  "mat-chu (='mất chữ', marker nomfoundation) → đổi thành <UNK>, KHÔNG xoá.",
  "Script strip + báo cáo số dòng đã sửa","1 giờ","Chưa",
  "mat-chu là chữ bị mất thật — xoá sẽ tạo câu thiếu nghĩa."],

 ["C3","P2","T3","Gazetteer",
  "Rebuild gazetteer trên corpus đã lọc blocklist",
  "Gazetteer cũ 13,188 entry build từ v1 (và bản đầu còn dùng dev labels = leakage). "
  "Nay có blocklist chính xác.",
  "Gazetteer mới từ 14,283 câu sạch; chỉ TITLE/ORG/LOC","4 giờ","Chưa",
  "Bạn đã kết luận đúng 03/08: ORG/TITLE/LOC dùng list, PER/DTM dùng rule."],

 ["C4","P2","T3","Gazetteer",
  "Rerun GazBertCRF với gazetteer sạch",
  "Thay số 0.8124. Phụ thuộc C3 + A4.",
  "GazBertCRF F1 trên v2-clean, gazetteer train-only","3 giờ","Chưa",
  "Chỉ làm nếu A4/A5 đã xong. Nice-to-have."],

 # ---------- P3: hoãn ----------
 ["D1","P3","Sau 30/08","Paper 2",
  "E3 (RAG) rerun trên v2","E3 cũ = 0.6888 trên v1. Phụ thuộc A5.",
  "E3 F1 trên v2","3 giờ","Hoãn","Chạy trên v1 là phí compute."],
 ["D2","P3","Sau 30/08","Paper 2",
  "E4 (seq1024) / E5 (rank32) / E6 (BGE-M3)","Config đã có trong configs/ nhưng chưa train.",
  "E4/E5/E6 metrics","10 giờ","Hoãn","Chỉ chạy sau khi E2 v2 có số ổn định."],
 ["D3","P3","Sau 30/08","Data",
  "Weak-label 105k ký tự chưa annotate",
  "~8,200 entity (+40%), gồm quyển 57-Lê Thái Tổ. KHÔNG giao test.",
  "Silver data + đánh giá tác động",
  "8 giờ","Hoãn",
  "RỦI RO CAO: 25–35% entity sẽ bị gán O sai (~2,400 nhãn sai). Cần partial annotation."],
 ["D4","P3","Sau 30/08","Data",
  "Bổ sung annotation section 57 (Lê Thái Tổ)",
  "1,068 câu, quyển LỚN NHẤT bộ sách, annotation không có một chữ.",
  "Annotation mới cho section 57","rất lớn","Hoãn",
  "Nếu không làm kịp → ghi vào Limitation."],
]

ws = sheet(wb, "Kế hoạch", HDR, TASKS,
           [6, 8, 15, 11, 34, 46, 40, 10, 11, 42], first=True)
for r in range(2, ws.max_row + 1):
    p = ws.cell(r, 2).value
    ws.cell(r, 2).fill = PF.get(p, P4)
    ws.cell(r, 2).alignment = CEN
    ws.cell(r, 2).font = BOLD
    if ws.cell(r, 9).value == "Đang làm":
        ws.cell(r, 9).fill = P2
ws.row_dimensions[1].height = 30
dv = DataValidation(type="list", formula1='"Chưa,Đang làm,Xong,Hoãn,Bỏ"', allow_blank=True)
ws.add_data_validation(dv)
dv.add(f"I2:I{ws.max_row}")

# ══════════════════════════════════════════════════ 2. KẾT QUẢ CẦN CHẠY LẠI
HDR2 = ["Kết quả cũ", "Giá trị", "Ngày", "Dữ liệu", "Còn dùng được?",
        "Lý do", "Task xử lý", "Output thay thế"]
RES = [
 ["CHisIEC GuwenBERT+CRF","0.9209","16/03/2026","CHisIEC","✅ GIỮ",
  "Dataset ngoại lai, không liên quan v1","—","Giữ nguyên, dùng làm điểm neo"],
 ["Paper 1 GuwenBERT-CRF","82.13%","01/06/2026","v1","❌ BỎ",
  "83.6% test trùng nguyên văn train → đo trí nhớ, không phải NER","A4","F1 mới trên v2-clean"],
 ["Paper 1 baseline (no CRF)","75.75%","01/06/2026","v1","❌ BỎ","cùng lý do","A4","F1 mới"],
 ["Paper 1 per-entity F1","PER .88 / LOC .85 / ORG .79 / DTM .79 / TITLE .76",
  "01/06/2026","v1","❌ BỎ","cùng lý do","A4","Bảng per-entity mới"],
 ["Thống kê corpus","7,301 câu · 1,291,822 token · 107,105 entity","01/06/2026","v1","❌ SAI SỐ",
  "Con số TỰ THÂN sai: nhân bản 3.7×. Thật ~1,961 unique → v2 còn 1,604","B2",
  "1,604 record · 20,724 entity (train+dev sau clean)"],
 ["DVSK baseline","0.6017","16/03/2026","v1","❌ BỎ","v1","A4","—"],
 ["DVSK transfer (CHisIEC→DVSK)","0.7575","16/03/2026","v1","❌ BỎ",
  "v1. Hướng transfer vẫn đúng, chỉ số sai","A4","Rerun transfer trên v2"],
 ["GazBertCRF (gaz train+dev)","0.8124","16/03/2026","v1 + leakage kép","❌ BỎ",
  "v1 CỘNG indirect leakage do dùng dev labels build gazetteer","C3+C4","GazBertCRF v2, gaz train-only"],
 ["GazBertCRF (gaz train-only)","đang chạy","13/04/2026","v1","❌ BỎ",
  "Đã sửa leakage gazetteer nhưng vẫn trên v1","C3+C4","—"],
 ["E1 (SFT epoch 1)","0.0952 (NaN)","27/07/2026","v1","❌ HUỶ",
  "NaN loss + v1. Base model tag đúng nghĩa nhưng gold sai → F1 ~0","—","Không rerun, bỏ khỏi paper"],
 ["E2 (pretrain+SFT 3 epoch)","0.7507","27/07/2026","v1","❌ BỎ",
  "Học thuộc pattern nhiễu","A5","E2 F1 trên v2-clean"],
 ["E3 (RAG TF-IDF 2-shot)","0.6888","27/07/2026","v1","❌ BỎ","v1","D1","E3 v2 (hoãn sau 30/08)"],
 ["Phân tích câu dài","F1 0.4851 với câu >200 ký tự","27/07/2026","v1","❌ BỎ",
  "v1. Nhưng hiện tượng truncate là thật, cần đo lại","B3","Phân tích độ dài trên v2"],
 ["Error analysis mục 6.4","error_analysis.txt","08/06/2026","v1","❌ LÀM LẠI",
  "Đang mô tả lỗi memorization, không phải lỗi NER","B3","error_analysis_v2.txt"],
 ["Web app / API NER / UI","—","05/2026","—","✅ GIỮ",
  "Không phụ thuộc chất lượng nhãn","—","Chỉ cần nạp lại model mới sau A4/A5"],
 ["Pipeline code, notebook, Git workflow","—","10/07/2026","—","✅ GIỮ","Hạ tầng, không phải kết quả","—","—"],
 ["Related work, guideline v1.0","—","2025–2026","—","✅ GIỮ","Không phụ thuộc dữ liệu","—","guideline nâng v1.1 ở C1"],
]
ws2 = sheet(wb, "Kết quả cần chạy lại", HDR2, RES, [30, 34, 12, 15, 13, 44, 11, 36])
for r in range(2, ws2.max_row + 1):
    v = ws2.cell(r, 5).value
    ws2.cell(r, 5).fill = OK if "GIỮ" in str(v) else BAD
    ws2.cell(r, 5).alignment = CEN

# ══════════════════════════════════════════════════ 3. ĐÃ XONG
HDR3 = ["Ngày", "Việc", "Kết quả / Output", "Ghi chú"]
DONE = [
 ["28–29/07/2026","Rebuild dataset v2 từ Excel gốc",
  "1,604 record · 23,088 entity · split 1284/160/160 · 0 leakage nội bộ · commit e6db6b8",
  "Đã push GitHub + upload Kaggle"],
 ["29/07/2026","Guideline annotation v1.0","docs/annotation_guideline.md · commit ca1ab37",
  "4 quyết định đã chốt: miếu hiệu=PER · tước=TITLE · generic title giữ TITLE · triều đại=ORG"],
 ["29/07/2026","Sinh 982 đề xuất rà gold test","D:\\bio_source\\gold_review_sheet.xlsx",
  "429 ADD · 182 RETYPE · 176 DEL · 158 BOUNDARY · 32 NOTE · 5 FIXTEXT"],
 ["03/08/2026","Crawl trọn DVSKTT từ nomfoundation.org",
  "19,909 câu unique · ~370K chữ Hán · 73/73 slug",
  "Đã verify với index site: các ID vắng KHÔNG tồn tại trên site"],
 ["03/08/2026","Khảo sát chiến lược gazetteer","docs/gazetteer_findings.md",
  "Xác nhận kết luận của bạn: ORG/TITLE/LOC dùng list · PER/DTM dùng rule"],
 ["04/08/2026","Clean nhãn train+dev","1,499 sửa · nhiễu train 5.2%→1.82% · dev 2.2%→0.17%",
  "scripts/clean_labels.py · BIO hợp lệ 0 lỗi/237,842 token · TEST không đụng"],
 ["04/08/2026","CHỨNG MINH text v2 lấy từ nomfoundation.org",
  "marker mat-chu + folio marker giống hệt · crawl phủ 83% ký tự v2",
  "⇒ crawl và gold KHÔNG độc lập"],
 ["04/08/2026","Map record → vị trí gốc",
  "99%+ record có section+position · record_source_map.json",
  "scripts/map_records_to_source.py"],
 ["04/08/2026","Định lượng leakage",
  "test SẠCH (1/160 chỉ dính boilerplate) · dev 6/160 leakage thật · blocklist 2,038 câu",
  "Còn 14,283/16,321 câu (87.5%) mine được"],
 ["04/08/2026","Phát hiện v2 chỉ phủ 71.6% DVSKTT",
  "10 section vắng hẳn, lớn nhất 57-Lê Thái Tổ (1,068 câu)",
  "PHẢI ghi vào Limitation của paper"],
]
sheet(wb, "Đã xong", HDR3, DONE, [14, 38, 62, 52])

# ══════════════════════════════════════════════════ 4. RỦI RO
HDR4 = ["Mức", "Rủi ro", "Ảnh hưởng", "Cách chặn", "Task"]
RISK = [
 ["CAO","gold_review_sheet.xlsx nằm ngoài git, không backup",
  "Mất toàn bộ công duyệt tay (58 dòng + phần đang làm)","Commit vào data/review/ ngay","A3"],
 ["CAO","Abstract đã đăng ký có thể quảng cáo quy mô corpus v1",
  "7,301 câu / 107,105 entity tự thân đã sai → vấn đề mô tả dữ liệu, không chỉ kết quả",
  "Rà abstract, xác nhận venue cho sửa trước 30/08","B1"],
 ["CAO","Chỉ còn 26 ngày tới 30/08",
  "Không kịp rerun + viết","Bám đúng P0→P1, hoãn hết P2/P3","—"],
 ["TRUNG BÌNH","Test chưa áp guideline trong khi train/dev đã áp",
  "Model bị trừ điểm vì trả lời ĐÚNG → F1 thấp giả tạo, lệch không đều giữa type",
  "Chạy clean_labels.py lên test","A1"],
 ["TRUNG BÌNH","F1 mới sẽ TỤT MẠNH so với 82.13%",
  "Nhìn như thất bại nếu không đóng khung đúng",
  "Đóng khung là contribution: phát hiện + sửa corpus + benchmark hiệu chỉnh","B5"],
 ["TRUNG BÌNH","review_rules.py đã mất khỏi máy",
  "Không chạy lại rule engine trên dev/train được","Chỉ còn output json; viết lại nếu cần","—"],
 ["THẤP","Metric 'nhãn thiểu số' mù với lỗi 1-lần và lỗi biên",
  "dev 0.17% nhìn đẹp nhưng vẫn sót lỗi (audit ngoài đã chỉ ra)",
  "Không dùng nó làm bằng chứng duy nhất","—"],
 ["THẤP","Silver-label 105k gây false negative 25–35%",
  "~2,400 nhãn sai hướng phủ định, nhiều hơn 1,499 lỗi vừa sửa",
  "Partial annotation / lọc câu / dùng làm feature","D3"],
]
ws4 = sheet(wb, "Rủi ro", HDR4, RISK, [12, 44, 50, 44, 8])
for r in range(2, ws4.max_row + 1):
    v = ws4.cell(r, 1).value
    ws4.cell(r, 1).fill = {"CAO": P1, "TRUNG BÌNH": P2}.get(v, P3)
    ws4.cell(r, 1).alignment = CEN
    ws4.cell(r, 1).font = BOLD

wb.save(OUT)
print("Đã ghi:", OUT)
for w in wb.worksheets:
    print(f"  {w.title:24s} {w.max_row-1} dòng")
