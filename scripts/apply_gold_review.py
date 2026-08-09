"""Apply duyệt gold review (NER_gold_review_sheet - Done.xlsx) vào data/raw/ner_sft/test.jsonl.

Đọc sheet Review (982 dòng), áp các dòng rõ ràng (OK + SỬA đã diễn giải được),
bỏ qua BỎ / chưa duyệt / SỬA mơ hồ (giữ nguyên).
"""
import json
import re
import sys
from pathlib import Path

import openpyxl

REVIEW_XLSX = r"C:\Users\PC\Downloads\NER_gold_review_sheet - Done.xlsx"
TEST_RAW = Path("data/raw/ner_sft/test.jsonl")
TEST_PROCESSED = Path("data/processed/ner_clean/test.jsonl")
TEST_BIO = Path("data/processed/ner_clean/test.bio.txt")

TAG_RE = re.compile(r"\{([^{}|]+)\|([A-Za-z]+)\}")

# (record, action, cur, prop) -> resolved override.
# op kinds: ('skip',) | ('retype', new_type) | ('boundary', new_text, new_type)
#           | ('add', new_text, new_type) | ('del',) | ('split', [(text,type), ...])
OVERRIDES = {
    (3, "DEL", "每年|LOC"): ("retype", "DTM"),
    (3, "DEL", "各司|ORG"): ("skip",),
    (28, "DEL", "官家|TITLE"): ("boundary", "醫官家", "TITLE"),
    (142, "DEL", "占奴|TITLE"): ("retype", "PER"),
    (147, "DEL", "大臣|LOC"): ("retype", "TITLE"),
    (1, "ADD", "節制府|ORG"): ("add", "節制府", "TITLE"),
    (140, "ADD", "御史|ORG"): ("add", "御史", "TITLE"),
    (44, "NOTE", "後江|LOC"): ("del",),
    (77, "NOTE", "開國功臣|TITLE"): ("del",),
    (91, "NOTE", "徽人|LOC"): ("del",),
    (95, "NOTE", "上京|LOC"): ("del",),
    (116, "NOTE", "明義土官|ORG"): ("split", [("明義", "LOC"), ("土官", "TITLE")]),
    (75, "BOUNDARY", "元方谷|PER"): ("split", [("元", "DTM"), ("方谷珍", "PER")]),
}
# Mơ hồ -> giữ nguyên (ghi lại để báo cáo, không áp)
AMBIGUOUS_SKIP = {
    (159, "DEL", "文官|ORG"),
    (118, "DEL", "北冦|ORG"),
    (1, "ADD", "太原|LOC"),
    (87, "ADD", "安󰘊|LOC"),
    (117, "ADD", "安󰘊|LOC"),
    (45, "NOTE", "莫慶王"),
    (64, "NOTE", None),  # cluster phức tạp, khớp theo record+action+cur startswith
    (127, "NOTE", "吏部上書"),
}


def parse_output(output):
    """output -> (plain_text, entities[(start,end,type,text)])"""
    plain = []
    entities = []
    pos = 0
    i = 0
    for m in re.finditer(r"\{[^{}]*\}|[^{}]", output):
        tok = m.group(0)
        tm = TAG_RE.fullmatch(tok)
        if tm:
            text, typ = tm.group(1), tm.group(2)
            entities.append([pos, pos + len(text), typ, text])
            plain.append(text)
            pos += len(text)
        else:
            plain.append(tok)
            pos += len(tok)
    return "".join(plain), entities


def find_ctx_span(plain, ctx):
    """ctx has 【target】 marking span. Return (start,end) in plain, or None."""
    if not ctx or "【" not in ctx or "】" not in ctx:
        return None
    pre, rest = ctx.split("【", 1)
    target, post = rest.split("】", 1)
    full = pre + target + post
    idx = plain.find(full)
    if idx == -1:
        # fallback: search without full context, just pre+target
        idx = plain.find(pre + target)
        if idx == -1:
            return None
    start = idx + len(pre)
    end = start + len(target)
    return start, end


def align_start(old_start, old_text, new_text):
    """old_text located at old_start; return start for new_text covering
    the same real-world span (handles prefix/suffix add or drop)."""
    if old_text == new_text:
        return old_start
    if old_text in new_text:
        return old_start - new_text.index(old_text)
    if new_text in old_text:
        return old_start + old_text.index(new_text)
    return old_start


def find_text_span(plain, text, used):
    """Locate next unused occurrence of `text` in plain. used: set of (start,end)."""
    pos = 0
    while True:
        idx = plain.find(text, pos)
        if idx == -1:
            return None
        span = (idx, idx + len(text))
        if span not in used:
            used.add(span)
            return span
        pos = idx + 1


def main():
    wb = openpyxl.load_workbook(REVIEW_XLSX, data_only=True)
    ws = wb["Review"]
    col = {c.value: i for i, c in enumerate(ws[1])}
    rows = list(ws.iter_rows(min_row=2, values_only=True))

    records = [json.loads(l) for l in open(TEST_RAW, encoding="utf-8")]
    parsed = []
    for rec in records:
        plain, ents = parse_output(rec["output"])
        assert plain == rec["input"], "plain/input mismatch"
        parsed.append({"input": rec["input"], "instruction": rec["instruction"], "entities": ents})

    applied, skipped_ambiguous, skipped_unreviewed, not_found = 0, 0, 0, []

    for r in rows:
        rec_i = r[col["record"]]
        action = r[col["action"]]
        duyet = (r[col["DUYỆT"]] or "").strip() if r[col["DUYỆT"]] is not None else ""
        cur = r[col["hiện tại"]]
        prop = r[col["đề xuất"]]
        ctx = r[col["ngữ cảnh"]]

        if duyet == "BỎ" or duyet == "":
            skipped_unreviewed += 1
            continue
        if action in ("NOTE", "FIXTEXT") and duyet == "OK" and action == "NOTE":
            continue  # NOTE OK = no-op by definition (cur==prop, informational)
        if action == "FIXTEXT":
            continue  # xử lý riêng thủ công (record156), bỏ qua trong pass tự động

        key = (rec_i, action, cur)
        override = OVERRIDES.get(key)

        if duyet == "SỬA" and override is None:
            skipped_ambiguous += 1
            continue

        rp = parsed[rec_i]
        plain = rp["input"]
        ents = rp["entities"]

        if override:
            kind = override[0]
            if kind == "skip":
                continue
            if kind == "retype":
                text = cur.split("|")[0]
                found = False
                for e in ents:
                    if e[3] == text and not found:
                        e[2] = override[1]
                        found = True
                if not found:
                    not_found.append((rec_i, action, cur, "retype-target-missing"))
                applied += 1
                continue
            if kind == "del":
                text = cur.split("|")[0]
                for e in list(ents):
                    if e[3] == text:
                        ents.remove(e)
                        break
                applied += 1
                continue
            if kind in ("boundary", "add"):
                new_text, new_type = override[1], override[2]
                span = find_ctx_span(plain, ctx) if ctx else None
                if span is not None:
                    start, end = span
                else:
                    # locate via old text (boundary) as anchor, then align
                    # start to the real-world position of new_text.
                    old_text = cur.split("|")[0] if cur and cur != "(chưa tag)" else None
                    if old_text and old_text in plain:
                        idx = plain.find(old_text)
                        start = align_start(idx, old_text, new_text)
                        end = start + len(new_text)
                    else:
                        start = end = None
                if start is None:
                    not_found.append((rec_i, action, cur, "add/boundary span not found"))
                    continue
                # remove overlapping existing entity
                for e in list(ents):
                    if e[0] < end and e[1] > start:
                        ents.remove(e)
                ents.append([start, start + len(new_text), new_type, new_text])
                applied += 1
                continue
            if kind == "split":
                anchor = find_ctx_span(plain, ctx) if ctx else None
                old_text = cur.split("|")[0]
                if anchor is None:
                    idx = plain.find(old_text)
                    anchor = (idx, idx + len(old_text)) if idx != -1 else None
                if anchor is None:
                    not_found.append((rec_i, action, cur, "split anchor not found"))
                    continue
                start, _ = anchor
                for e in list(ents):
                    if e[3] == old_text:
                        ents.remove(e)
                        break
                offset = start
                for text, typ in override[1]:
                    ents.append([offset, offset + len(text), typ, text])
                    offset += len(text)
                applied += 1
                continue

        # Không override -> áp trực tiếp theo action (OK rows)
        if action == "DEL":
            text = cur.split("|")[0]
            removed = False
            for e in list(ents):
                if e[3] == text:
                    ents.remove(e)
                    removed = True
                    break
            if not removed:
                not_found.append((rec_i, action, cur, "del-target-missing"))
            else:
                applied += 1
            continue

        if action in ("ADD", "RETYPE", "BOUNDARY"):
            if "+" in prop or (cur and "+" in cur):
                # compound merge/split proposal without unambiguous single-span
                # locator: too risky to auto-apply, leave as-is.
                skipped_ambiguous += 1
                continue
            new_text, new_type = prop.rsplit("|", 1)
            span = find_ctx_span(plain, ctx)
            if span is None and action == "BOUNDARY" and not ctx:
                # BOUNDARY without ngữ cảnh: no reliable anchor for a span
                # that changes length/position. Too risky to auto-apply.
                skipped_ambiguous += 1
                continue
            if span is None:
                used = rp.setdefault("_used", set())
                locate_text = cur.split("|")[0] if cur and cur != "(chưa tag)" else new_text
                span = find_text_span(plain, locate_text, used)
                if span is not None and locate_text != new_text:
                    start0, _ = span
                    aligned = align_start(start0, locate_text, new_text)
                    span = (aligned, aligned + len(new_text))
            if span is None:
                not_found.append((rec_i, action, cur, prop, "span not found"))
                continue
            start, end = span
            for e in list(ents):
                if e[0] < end and e[1] > start:
                    ents.remove(e)
            ents.append([start, start + len(new_text), new_type, new_text])
            applied += 1
            continue

    # Rebuild output for changed records
    out_records = []
    for rp, orig in zip(parsed, records):
        ents = sorted(rp["entities"], key=lambda e: e[0])
        plain = rp["input"]
        pieces = []
        pos = 0
        for s, e, typ, text in ents:
            if s < pos:
                continue  # overlap guard, skip
            pieces.append(plain[pos:s])
            pieces.append("{" + text + "|" + typ + "}")
            pos = e
        pieces.append(plain[pos:])
        new_output = "".join(pieces)
        out_records.append({"instruction": orig["instruction"], "input": plain, "output": new_output})

    n_changed = sum(1 for a, b in zip(records, out_records) if a["output"] != b["output"])

    print(f"applied ops: {applied}")
    print(f"skipped (BỎ/chưa duyệt): {skipped_unreviewed}")
    print(f"skipped (SỬA mơ hồ, giữ nguyên): {skipped_ambiguous}")
    print(f"not_found ({len(not_found)}):")
    for x in not_found:
        print("  ", x)
    print(f"records changed: {n_changed} / {len(records)}")

    with open(TEST_RAW, "w", encoding="utf-8") as f:
        for r in out_records:
            f.write(json.dumps(r, ensure_ascii=False) + "\n")
    with open(TEST_PROCESSED, "w", encoding="utf-8") as f:
        for r in out_records:
            f.write(json.dumps(r, ensure_ascii=False) + "\n")

    print("wrote", TEST_RAW, "and", TEST_PROCESSED)


if __name__ == "__main__":
    main()
